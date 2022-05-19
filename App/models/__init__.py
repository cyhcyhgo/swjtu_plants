import openpyxl as op
import os

from .user import Users
from .plant import Plants, Info, Position, Picture


def initial_database(database):
    """从Excel读取数据，初始化数据库"""
    del_list = [Plants, Info, Position, Picture]
    for table in del_list:
        for member in table.query.all():
            database.session.delete(member)

    # 读取Excel
    d = op.load_workbook(os.path.abspath(os.path.dirname(os.path.dirname(__file__))) + r'\static\intro.xlsx')
    table = d['Sheet1']
    del d

    j = 19
    while table.cell(row=j, column=1) == '':
        j = j - 1

    i = 2
    while i <= j:
        p1 = Plants(id=int(i-1), plant_name=str(table.cell(row=i, column=2).value), time=str(table.cell(row=i, column=5).value),
                    language=str(table.cell(row=i, column=6).value), type=str(table.cell(row=i, column=10).value),
                    family=str(table.cell(row=i, column=11).value), season=str(table.cell(row=i, column=12).value))
        p2 = Info(id=int(i-1), academic=str(table.cell(row=i, column=3).value), literary=str(table.cell(row=i, column=4).value),
                  brief=str(table.cell(row=i, column=15).value), plants_id=i - 1)
        p3 = Position(id=int(i-1), longitude=float(table.cell(row=i, column=7).value),
                      latitude=float(table.cell(row=i, column=8).value),
                      area=str(table.cell(row=i, column=9).value), plants_id=i - 1)
        p4 = Picture(id=int(i-1), search_p=str(table.cell(row=i, column=13).value), detail_p=str(table.cell(row=i, column=14).value),
                     plants_id=i - 1)
        database.session.add(p1)
        database.session.add(p2)
        database.session.add(p3)
        database.session.add(p4)
        i = i + 1
    # for member in Users.query.all():
    #     database.session.delete(member)
    # p5 = Users(id=1, username='CYH', password='123456', isAdministrator=1, confirmed=True)
    # p6 = Users(id=2, username='GX', password='654321', isAdministrator=0, confirmed=True)
    # database.session.add(p5)
    # database.session.add(p6)
    database.session.commit()
